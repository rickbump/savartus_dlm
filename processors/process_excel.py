import os
from openpyxl import load_workbook


def process_excel(file_path):
    metadata = {}
    try:
        # Basic file metadata
        file_metadata = os.stat(file_path)
        metadata['file_size'] = file_metadata.st_size
        metadata['file_modified'] = file_metadata.st_mtime

        # Load workbook and extract basic properties
        workbook = load_workbook(filename=file_path, read_only=True)
        metadata['sheet_count'] = len(workbook.sheetnames)
        metadata['sheet_names'] = workbook.sheetnames

        # Workbook properties
        properties = workbook.properties
        metadata['author'] = properties.creator
        metadata['title'] = properties.title
        metadata['created'] = properties.created
        metadata['modified'] = properties.modified

        # Process metadata for each sheet
        for sheet_name in workbook.sheetnames:
            sheet_metadata = {}
            sheet = workbook[sheet_name]

            # Sheet-level metadata
            sheet_metadata['row_count'] = sheet.max_row
            sheet_metadata['column_count'] = sheet.max_column
            sheet_metadata['frozen_panes'] = sheet.freeze_panes
            sheet_metadata['merged_cells'] = list(sheet.merged_cells)
            sheet_metadata['protection'] = sheet.protection.sheet

            # Collect cell-level metadata for the first few cells (for illustration)
            cell_metadata = {}
            for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
                for cell in row:
                    cell_info = {
                        'value': cell.value,
                        'data_type': cell.data_type,
                        'style': cell.style,
                        'is_merged': cell.merged,
                        'formula': cell.formula if cell.has_formula else 'No formula'
                    }
                    cell_metadata[f'{cell.coordinate}'] = cell_info

            sheet_metadata['cells'] = cell_metadata
            metadata[sheet_name] = sheet_metadata

    except Exception as e:
        metadata['excel_error'] = str(e)

    return metadata